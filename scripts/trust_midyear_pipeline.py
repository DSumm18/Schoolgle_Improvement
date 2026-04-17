#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCHOOLS = ["CVPS", "CHPS", "FPS", "GHPS", "HPS", "LPS", "LGPS"]
YEAR_ORDER = ["EYFS", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5", "Year 6"]

YEAR_LAYOUTS: dict[str, dict[str, Any]] = {
    "EYFS": {
        "counts": {"cohort": "B", "send": "C", "ehcp": "D", "fsm": "E"},
        "groups": {
            "all": {"GLD": "I"},
            "fsm": {"GLD": "Q"},
            "non_fsm": {"GLD": "Y"},
        },
        "core_metric": "GLD",
    },
    "Year 1": {
        "counts": {"cohort": "B", "send": "C", "ehcp": "D", "fsm": "E"},
        "groups": {
            "all": {
                "R_ARE": "F",
                "R_GD": "G",
                "W_ARE": "H",
                "W_GD": "I",
                "M_ARE": "J",
                "M_GD": "K",
                "C_ARE": "L",
                "C_GD": "M",
                "PHONICS": "N",
            },
            "fsm": {
                "R_ARE": "O",
                "R_GD": "P",
                "W_ARE": "Q",
                "W_GD": "R",
                "M_ARE": "S",
                "M_GD": "T",
                "C_ARE": "U",
                "C_GD": "V",
                "PHONICS": "W",
            },
            "non_fsm": {
                "R_ARE": "X",
                "R_GD": "Y",
                "W_ARE": "Z",
                "W_GD": "AA",
                "M_ARE": "AB",
                "M_GD": "AC",
                "C_ARE": "AD",
                "C_GD": "AE",
                "PHONICS": "AF",
            },
        },
        "core_metric": "C_ARE",
    },
    "Year 2": {
        "counts": {"cohort": "B", "send": "C", "ehcp": "D", "fsm": "E"},
        "groups": {
            "all": {
                "R_ARE": "F",
                "R_GD": "G",
                "W_ARE": "H",
                "W_GD": "I",
                "M_ARE": "J",
                "M_GD": "K",
                "C_ARE": "L",
                "C_GD": "M",
                "PHONICS": "N",
            },
            "fsm": {
                "R_ARE": "O",
                "R_GD": "P",
                "W_ARE": "Q",
                "W_GD": "R",
                "M_ARE": "S",
                "M_GD": "T",
                "C_ARE": "U",
                "C_GD": "V",
                "PHONICS": "W",
            },
            "non_fsm": {
                "R_ARE": "X",
                "R_GD": "Y",
                "W_ARE": "Z",
                "W_GD": "AA",
                "M_ARE": "AB",
                "M_GD": "AC",
                "C_ARE": "AD",
                "C_GD": "AE",
                "PHONICS": "AF",
            },
        },
        "core_metric": "C_ARE",
    },
    "Year 3": {
        "counts": {"cohort": "B", "send": "C", "ehcp": "D", "fsm": "E"},
        "groups": {
            "all": {
                "R_ARE": "F",
                "R_GD": "G",
                "W_ARE": "H",
                "W_GD": "I",
                "M_ARE": "J",
                "M_GD": "K",
                "C_ARE": "L",
                "C_GD": "M",
            },
            "fsm": {
                "R_ARE": "N",
                "R_GD": "O",
                "W_ARE": "P",
                "W_GD": "Q",
                "M_ARE": "R",
                "M_GD": "S",
                "C_ARE": "T",
                "C_GD": "U",
            },
            "non_fsm": {
                "R_ARE": "V",
                "R_GD": "W",
                "W_ARE": "X",
                "W_GD": "Y",
                "M_ARE": "Z",
                "M_GD": "AA",
                "C_ARE": "AB",
                "C_GD": "AC",
            },
        },
        "core_metric": "C_ARE",
    },
    "Year 4": {
        "counts": {"cohort": "B", "send": "C", "ehcp": "D", "fsm": "E"},
        "groups": {
            "all": {
                "R_ARE": "F",
                "R_GD": "G",
                "W_ARE": "H",
                "W_GD": "I",
                "M_ARE": "J",
                "M_GD": "K",
                "C_ARE": "L",
                "C_GD": "M",
                "MTC": "N",
            },
            "fsm": {
                "R_ARE": "O",
                "R_GD": "P",
                "W_ARE": "Q",
                "W_GD": "R",
                "M_ARE": "S",
                "M_GD": "T",
                "C_ARE": "U",
                "C_GD": "V",
                "MTC": "W",
            },
            "non_fsm": {
                "R_ARE": "X",
                "R_GD": "Y",
                "W_ARE": "Z",
                "W_GD": "AA",
                "M_ARE": "AB",
                "M_GD": "AC",
                "C_ARE": "AD",
                "C_GD": "AE",
                "MTC": "AF",
            },
        },
        "core_metric": "C_ARE",
    },
    "Year 5": {
        "counts": {"cohort": "B", "send": "C", "ehcp": "D", "fsm": "E"},
        "groups": {
            "all": {
                "R_ARE": "F",
                "R_GD": "G",
                "W_ARE": "H",
                "W_GD": "I",
                "M_ARE": "J",
                "M_GD": "K",
                "C_ARE": "L",
                "C_GD": "M",
            },
            "fsm": {
                "R_ARE": "N",
                "R_GD": "O",
                "W_ARE": "P",
                "W_GD": "Q",
                "M_ARE": "R",
                "M_GD": "S",
                "C_ARE": "T",
                "C_GD": "U",
            },
            "non_fsm": {
                "R_ARE": "V",
                "R_GD": "W",
                "W_ARE": "X",
                "W_GD": "Y",
                "M_ARE": "Z",
                "M_GD": "AA",
                "C_ARE": "AB",
                "C_GD": "AC",
            },
        },
        "core_metric": "C_ARE",
    },
    "Year 6": {
        "counts": {"cohort": "B", "send": "C", "ehcp": "D", "fsm": "E"},
        "groups": {
            "all": {
                "R_ARE": "F",
                "R_GD": "G",
                "W_ARE": "H",
                "W_GD": "I",
                "M_ARE": "J",
                "M_GD": "K",
                "C_ARE": "L",
                "C_GD": "M",
            },
            "fsm": {
                "R_ARE": "N",
                "R_GD": "O",
                "W_ARE": "P",
                "W_GD": "Q",
                "M_ARE": "R",
                "M_GD": "S",
                "C_ARE": "T",
                "C_GD": "U",
            },
            "non_fsm": {
                "R_ARE": "V",
                "R_GD": "W",
                "W_ARE": "X",
                "W_GD": "Y",
                "M_ARE": "Z",
                "M_GD": "AA",
                "C_ARE": "AB",
                "C_GD": "AC",
            },
        },
        "core_metric": "C_ARE",
    },
}

DISPLAY_METRIC = {
    "EYFS": "GLD",
    "Year 1": "Combined ARE",
    "Year 2": "Combined ARE",
    "Year 3": "Combined ARE",
    "Year 4": "Combined ARE",
    "Year 5": "Combined ARE",
    "Year 6": "Combined ARE",
}


def parse_count(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"-?\d+(?:\.\d+)?", str(value).replace(",", ""))
    return float(match.group(0)) if match else None


def parse_rate(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        if value < 0:
            return None
        if value <= 1:
            return float(value)
        if value <= 100:
            return float(value) / 100
        return None

    text = str(value).strip().replace("%%", "%")
    if not text:
        return None

    pct_match = re.search(r"(-?\d+(?:\.\d+)?)\s*%", text)
    if pct_match:
        return float(pct_match.group(1)) / 100

    num_match = re.search(r"-?\d+(?:\.\d+)?", text)
    if num_match:
        number = float(num_match.group(0))
        if number <= 1:
            return number
        if number <= 100:
            return number / 100
    return None


def pct(value: float | None) -> str:
    if value is None:
        return "-"
    return f"{value * 100:.1f}%"


def round3(value: float | None) -> float | None:
    if value is None:
        return None
    return round(value, 3)


def extract_records(workbook_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, dict[str, Any]]]:
    wb = load_workbook(workbook_path, data_only=True)

    records: list[dict[str, Any]] = []
    raw_text_cells: list[dict[str, Any]] = []
    national_rows: dict[str, dict[str, Any]] = {}

    for year in YEAR_ORDER:
        ws = wb[year]
        layout = YEAR_LAYOUTS[year]

        # National row capture (for benchmark coverage checks)
        national_row = None
        for row in range(1, 30):
            first_cell = ws[f"A{row}"].value
            if isinstance(first_cell, str) and first_cell.startswith("National"):
                national_row = row
                break
        national_rows[year] = {
            "row": national_row,
            "label": ws[f"A{national_row}"].value if national_row else None,
            "values": {},
        }

        if national_row:
            for col in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
                value = ws[f"{col}{national_row}"].value
                if value not in (None, ""):
                    national_rows[year]["values"][col] = value
            for col in [
                "AA",
                "AB",
                "AC",
                "AD",
                "AE",
                "AF",
                "AG",
                "AH",
                "AI",
                "AJ",
            ]:
                value = ws[f"{col}{national_row}"].value
                if value not in (None, ""):
                    national_rows[year]["values"][col] = value

        for row in range(1, 30):
            school = ws[f"A{row}"].value
            if school not in SCHOOLS:
                continue

            record: dict[str, Any] = {"year": year, "school": school, "row": row}

            for field, col in layout["counts"].items():
                raw_value = ws[f"{col}{row}"].value
                record[field] = parse_count(raw_value)
                if isinstance(raw_value, str) and raw_value.strip():
                    raw_text_cells.append(
                        {
                            "year": year,
                            "school": school,
                            "field": field,
                            "column": col,
                            "raw": raw_value,
                        }
                    )

            for group, metrics in layout["groups"].items():
                for metric, col in metrics.items():
                    raw_value = ws[f"{col}{row}"].value
                    field_name = f"{group}_{metric}"
                    record[field_name] = parse_rate(raw_value)
                    if isinstance(raw_value, str) and raw_value.strip():
                        raw_text_cells.append(
                            {
                                "year": year,
                                "school": school,
                                "field": field_name,
                                "column": col,
                                "raw": raw_value,
                            }
                        )

            records.append(record)

    return records, raw_text_cells, national_rows


def weighted_average(values: list[tuple[float, float]]) -> float | None:
    if not values:
        return None
    denominator = sum(weight for weight, _ in values)
    if denominator == 0:
        return None
    return sum(weight * value for weight, value in values) / denominator


def run_checks(records: list[dict[str, Any]], raw_text_cells: list[dict[str, Any]], national_rows: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []

    by_year_school = {(item["year"], item["school"]): item for item in records}

    for year in YEAR_ORDER:
        layout = YEAR_LAYOUTS[year]
        core_metric = layout["core_metric"]

        # National benchmark coverage
        national = national_rows[year]
        if national["row"] is None:
            issues.append(
                {
                    "severity": "warning",
                    "year": year,
                    "school": "ALL",
                    "check": "benchmark_missing",
                    "message": f"No national benchmark row present for {year}.",
                }
            )
        elif year == "Year 6":
            weird_columns = []
            for col, value in national["values"].items():
                if col in {"V", "X", "Z"} and isinstance(value, (int, float)) and value > 1:
                    weird_columns.append((col, value))
            if weird_columns:
                issues.append(
                    {
                        "severity": "warning",
                        "year": year,
                        "school": "ALL",
                        "check": "benchmark_format",
                        "message": "Year 6 national Not-FSM comparators appear in whole percentages while other benchmark fields use decimals.",
                        "details": weird_columns,
                    }
                )

        for school in SCHOOLS:
            record = by_year_school.get((year, school))
            if record is None:
                issues.append(
                    {
                        "severity": "critical",
                        "year": year,
                        "school": school,
                        "check": "missing_row",
                        "message": f"Missing row for {school} in {year}.",
                    }
                )
                continue

            cohort = record.get("cohort")
            send = record.get("send")
            ehcp = record.get("ehcp")
            fsm = record.get("fsm")

            # Integer and hierarchy checks
            for field in ["cohort", "send", "ehcp", "fsm"]:
                value = record.get(field)
                if value is None:
                    continue
                if value < 0:
                    issues.append(
                        {
                            "severity": "critical",
                            "year": year,
                            "school": school,
                            "check": "negative_count",
                            "message": f"{field.upper()} is negative ({value}).",
                        }
                    )
                if abs(value - round(value)) > 1e-6:
                    issues.append(
                        {
                            "severity": "critical",
                            "year": year,
                            "school": school,
                            "check": "non_integer_count",
                            "message": f"{field.upper()} is non-integer ({value}).",
                        }
                    )

            if cohort is not None and send is not None and send > cohort + 1e-9:
                issues.append(
                    {
                        "severity": "critical",
                        "year": year,
                        "school": school,
                        "check": "send_gt_cohort",
                        "message": f"SEND ({send}) is greater than cohort ({cohort}).",
                    }
                )
            if send is not None and ehcp is not None and ehcp > send + 1e-9:
                issues.append(
                    {
                        "severity": "critical",
                        "year": year,
                        "school": school,
                        "check": "ehcp_gt_send",
                        "message": f"EHCP ({ehcp}) is greater than SEND ({send}).",
                    }
                )
            if cohort is not None and fsm is not None and fsm > cohort + 1e-9:
                issues.append(
                    {
                        "severity": "critical",
                        "year": year,
                        "school": school,
                        "check": "fsm_gt_cohort",
                        "message": f"FSM ({fsm}) is greater than cohort ({cohort}).",
                    }
                )

            # Missing all-pupil core metric
            all_core = record.get(f"all_{core_metric}")
            if all_core is None:
                issues.append(
                    {
                        "severity": "critical",
                        "year": year,
                        "school": school,
                        "check": "missing_core_metric",
                        "message": f"All-pupil {core_metric} is missing.",
                    }
                )

            # Subgroup missingness where denominator exists
            if cohort is not None and fsm is not None:
                non_fsm = cohort - fsm
                fsm_core = record.get(f"fsm_{core_metric}")
                non_fsm_core = record.get(f"non_fsm_{core_metric}")

                if fsm > 0 and fsm_core is None:
                    issues.append(
                        {
                            "severity": "warning",
                            "year": year,
                            "school": school,
                            "check": "missing_fsm_subgroup",
                            "message": f"FSM subgroup {core_metric} missing despite FSM count {fsm:.0f}.",
                        }
                    )
                if non_fsm > 0 and non_fsm_core is None:
                    issues.append(
                        {
                            "severity": "warning",
                            "year": year,
                            "school": school,
                            "check": "missing_nonfsm_subgroup",
                            "message": f"Non-FSM subgroup {core_metric} missing despite non-FSM count {non_fsm:.0f}.",
                        }
                    )

                # Small subgroup reliability warnings
                if 0 < fsm < 5:
                    issues.append(
                        {
                            "severity": "info",
                            "year": year,
                            "school": school,
                            "check": "small_fsm",
                            "message": f"FSM subgroup very small (n={fsm:.0f}); subgroup percentages may be volatile.",
                        }
                    )
                if 0 < non_fsm < 10:
                    issues.append(
                        {
                            "severity": "info",
                            "year": year,
                            "school": school,
                            "check": "small_nonfsm",
                            "message": f"Non-FSM subgroup very small (n={non_fsm:.0f}); subgroup percentages may be volatile.",
                        }
                    )

            # Attainment logic checks for Years 1-6
            if year != "EYFS":
                for group in ["all", "fsm", "non_fsm"]:
                    for subject in ["R", "W", "M", "C"]:
                        are = record.get(f"{group}_{subject}_ARE")
                        gd = record.get(f"{group}_{subject}_GD")
                        if are is not None and gd is not None and gd > are + 0.005:
                            issues.append(
                                {
                                    "severity": "critical",
                                    "year": year,
                                    "school": school,
                                    "check": "gd_gt_are",
                                    "message": f"{group.upper()} {subject}_GD ({pct(gd)}) exceeds {subject}_ARE ({pct(are)}).",
                                }
                            )

                    c_are = record.get(f"{group}_C_ARE")
                    r_are = record.get(f"{group}_R_ARE")
                    w_are = record.get(f"{group}_W_ARE")
                    m_are = record.get(f"{group}_M_ARE")
                    if None not in (c_are, r_are, w_are, m_are):
                        min_core = min(r_are, w_are, m_are)
                        if c_are > min_core + 0.005:
                            issues.append(
                                {
                                    "severity": "warning",
                                    "year": year,
                                    "school": school,
                                    "check": "combined_gt_component",
                                    "message": f"{group.upper()} C_ARE ({pct(c_are)}) exceeds min(R/W/M ARE {pct(min_core)}).",
                                }
                            )

            # Weighted consistency checks
            if cohort is not None and fsm is not None and cohort > 0:
                non_fsm = cohort - fsm
                if non_fsm >= 0:
                    for metric in ["GLD", "R_ARE", "R_GD", "W_ARE", "W_GD", "M_ARE", "M_GD", "C_ARE", "C_GD", "PHONICS", "MTC"]:
                        all_value = record.get(f"all_{metric}")
                        fsm_value = record.get(f"fsm_{metric}")
                        non_fsm_value = record.get(f"non_fsm_{metric}")
                        if None in (all_value, fsm_value, non_fsm_value):
                            continue
                        expected = ((fsm * fsm_value) + (non_fsm * non_fsm_value)) / cohort
                        diff = all_value - expected
                        abs_diff = abs(diff)
                        if abs_diff > 0.10:
                            severity = "critical"
                        elif abs_diff > 0.05:
                            severity = "warning"
                        else:
                            continue
                        issues.append(
                            {
                                "severity": severity,
                                "year": year,
                                "school": school,
                                "check": "weighted_mismatch",
                                "message": f"{metric} all-pupil value ({pct(all_value)}) differs from weighted subgroup expectation ({pct(expected)}) by {diff * 100:+.1f}pp.",
                            }
                        )

    # Raw text entries are not all bad, but highlight for review
    for item in raw_text_cells:
        issues.append(
            {
                "severity": "info",
                "year": item["year"],
                "school": item["school"],
                "check": "text_entry",
                "message": f"Non-standard text entry in {item['field']} ({item['column']}): {item['raw']!r}.",
            }
        )

    severity_rank = {"critical": 0, "warning": 1, "info": 2}
    issues.sort(key=lambda item: (severity_rank.get(item["severity"], 9), YEAR_ORDER.index(item["year"]), item["school"], item["check"]))
    return issues


def build_summary(records: list[dict[str, Any]], issues: list[dict[str, Any]]) -> dict[str, Any]:
    by_year: dict[str, list[dict[str, Any]]] = defaultdict(list)
    by_school: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for record in records:
        by_year[record["year"]].append(record)
        by_school[record["school"]].append(record)

    trust_outcome_series = []
    trust_gap_series = []

    for year in YEAR_ORDER:
        core_metric = YEAR_LAYOUTS[year]["core_metric"]
        rows = by_year[year]

        outcome_pairs: list[tuple[float, float]] = []
        fsm_pairs: list[tuple[float, float]] = []
        non_fsm_pairs: list[tuple[float, float]] = []

        for row in rows:
            cohort = row.get("cohort")
            fsm = row.get("fsm")
            if cohort is not None:
                core_value = row.get(f"all_{core_metric}")
                if core_value is not None:
                    outcome_pairs.append((cohort, core_value))

            if None not in (cohort, fsm):
                non_fsm = cohort - fsm
                fsm_value = row.get(f"fsm_{core_metric}")
                non_fsm_value = row.get(f"non_fsm_{core_metric}")
                if fsm > 0 and fsm_value is not None:
                    fsm_pairs.append((fsm, fsm_value))
                if non_fsm > 0 and non_fsm_value is not None:
                    non_fsm_pairs.append((non_fsm, non_fsm_value))

        trust_value = weighted_average(outcome_pairs)
        trust_fsm = weighted_average(fsm_pairs)
        trust_non_fsm = weighted_average(non_fsm_pairs)
        gap_value = None
        if trust_fsm is not None and trust_non_fsm is not None:
            gap_value = trust_non_fsm - trust_fsm

        trust_outcome_series.append(
            {
                "year": year,
                "value": round3(trust_value),
                "metric": DISPLAY_METRIC[year],
            }
        )
        trust_gap_series.append(
            {
                "year": year,
                "value": round3(gap_value),
                "metric": DISPLAY_METRIC[year],
            }
        )

    school_scores = []
    for school in SCHOOLS:
        weighted_values: list[tuple[float, float]] = []
        for row in by_school[school]:
            year = row["year"]
            core_metric = YEAR_LAYOUTS[year]["core_metric"]
            value = row.get(f"all_{core_metric}")
            cohort = row.get("cohort")
            if value is None or cohort is None:
                continue
            weighted_values.append((cohort, value))

        school_scores.append(
            {
                "school": school,
                "value": round3(weighted_average(weighted_values)),
            }
        )

    school_scores.sort(key=lambda item: item["value"] if item["value"] is not None else -999, reverse=True)

    heatmap = []
    for school in SCHOOLS:
        row = {"school": school}
        for year in YEAR_ORDER:
            record = next((item for item in by_school[school] if item["year"] == year), None)
            if record is None:
                row[year] = None
                continue
            core_metric = YEAR_LAYOUTS[year]["core_metric"]
            row[year] = round3(record.get(f"all_{core_metric}"))
        heatmap.append(row)

    issue_counts = {
        "critical": sum(1 for issue in issues if issue["severity"] == "critical"),
        "warning": sum(1 for issue in issues if issue["severity"] == "warning"),
        "info": sum(1 for issue in issues if issue["severity"] == "info"),
    }

    trust_valid = [item for item in trust_outcome_series if item["value"] is not None]
    strongest_year = max(trust_valid, key=lambda item: item["value"]) if trust_valid else None
    weakest_year = min(trust_valid, key=lambda item: item["value"]) if trust_valid else None

    gap_valid = [item for item in trust_gap_series if item["value"] is not None]
    widest_gap = max(gap_valid, key=lambda item: item["value"]) if gap_valid else None

    return {
        "trust_outcome_series": trust_outcome_series,
        "trust_gap_series": trust_gap_series,
        "school_scores": school_scores,
        "heatmap": heatmap,
        "issue_counts": issue_counts,
        "strongest_year": strongest_year,
        "weakest_year": weakest_year,
        "widest_gap": widest_gap,
    }


def write_csv(records: list[dict[str, Any]], path: Path) -> None:
    fields = ["year", "school", "row", "cohort", "send", "ehcp", "fsm"]
    all_keys: set[str] = set()
    for record in records:
        all_keys.update(record.keys())
    dynamic_fields = sorted([name for name in all_keys if name not in fields])
    all_fields = fields + dynamic_fields

    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=all_fields)
        writer.writeheader()
        for record in records:
            writer.writerow(record)


def build_dashboard_html(summary: dict[str, Any], issues: list[dict[str, Any]]) -> str:
    severe_issues = [issue for issue in issues if issue["severity"] in {"critical", "warning"}]
    top_issues = severe_issues[:28]

    payload = {
        "trustOutcome": summary["trust_outcome_series"],
        "trustGap": summary["trust_gap_series"],
        "schoolScores": summary["school_scores"],
        "heatmap": summary["heatmap"],
        "issues": issues,
        "topIssues": top_issues,
        "issueCounts": summary["issue_counts"],
        "strongestYear": summary["strongest_year"],
        "weakestYear": summary["weakest_year"],
        "widestGap": summary["widest_gap"],
    }

    html_template = """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Trust Mid-Year Insights Dashboard</title>
  <style>
    :root {
      --bg: #070f1d;
      --bg2: #0f223b;
      --panel: rgba(14, 31, 57, 0.86);
      --panel-soft: rgba(23, 49, 86, 0.72);
      --border: rgba(148, 189, 255, 0.28);
      --text: #edf5ff;
      --muted: #a2bee3;
      --accent: #2dd4bf;
      --accent2: #60a5fa;
      --warn: #fbbf24;
      --bad: #fb7185;
      --good: #4ade80;
      --ink: #031024;
      --bar-bg: rgba(255, 255, 255, 0.11);
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: "Avenir Next", "Segoe UI", "Helvetica Neue", sans-serif;
      color: var(--text);
      background:
        radial-gradient(1000px 500px at 8% -10%, rgba(56, 189, 248, 0.33), transparent 65%),
        radial-gradient(900px 420px at 88% -10%, rgba(99, 102, 241, 0.29), transparent 65%),
        linear-gradient(165deg, var(--bg), var(--bg2));
      min-height: 100vh;
    }
    .wrap {
      width: min(1400px, 95vw);
      margin: 22px auto 44px;
    }
    .hero {
      background: linear-gradient(140deg, rgba(30, 78, 145, 0.5), rgba(18, 56, 109, 0.42));
      border: 1px solid var(--border);
      border-radius: 18px;
      padding: 22px 24px;
      box-shadow: 0 22px 60px rgba(0, 0, 0, 0.35);
      position: relative;
      overflow: hidden;
    }
    .hero::after {
      content: "";
      position: absolute;
      inset: -140px -120px auto auto;
      width: 300px;
      height: 300px;
      background: radial-gradient(circle at center, rgba(45, 212, 191, 0.35), transparent 70%);
      pointer-events: none;
    }
    h1 {
      margin: 0;
      font-size: 31px;
      letter-spacing: 0.2px;
    }
    .subtitle {
      margin: 8px 0 0;
      color: var(--muted);
      font-size: 14px;
      max-width: 1000px;
    }
    .kpis {
      margin-top: 16px;
      display: grid;
      grid-template-columns: repeat(5, minmax(0, 1fr));
      gap: 10px;
    }
    .kpi {
      background: var(--panel-soft);
      border: 1px solid var(--border);
      border-radius: 13px;
      padding: 11px 12px;
      min-height: 86px;
      transition: transform 0.28s ease;
    }
    .kpi:hover {
      transform: translateY(-2px);
    }
    .kpi .label {
      color: var(--muted);
      font-size: 11px;
      text-transform: uppercase;
      letter-spacing: 1px;
    }
    .kpi .value {
      margin-top: 6px;
      font-size: 24px;
      font-weight: 700;
      line-height: 1.1;
    }
    .kpi .sub {
      margin-top: 6px;
      color: #c2d8f5;
      font-size: 12px;
    }

    .layout {
      margin-top: 14px;
      display: grid;
      grid-template-columns: 1.15fr 0.85fr;
      gap: 12px;
    }
    .stack {
      display: grid;
      gap: 12px;
    }
    .panel {
      background: var(--panel);
      border: 1px solid var(--border);
      border-radius: 14px;
      padding: 14px;
      box-shadow: 0 16px 44px rgba(0, 0, 0, 0.22);
    }
    .panel h2 {
      margin: 0;
      font-size: 18px;
      letter-spacing: 0.1px;
    }
    .panel p {
      margin: 6px 0 0;
      color: var(--muted);
      font-size: 12px;
    }
    .panel-top {
      display: flex;
      justify-content: space-between;
      align-items: flex-start;
      gap: 10px;
      margin-bottom: 8px;
    }
    .control {
      border: 1px solid var(--border);
      background: rgba(13, 32, 60, 0.78);
      color: var(--text);
      border-radius: 8px;
      padding: 5px 9px;
      font-size: 12px;
    }
    .chart {
      width: 100%;
      height: 250px;
      background: linear-gradient(180deg, rgba(12, 26, 50, 0.6), rgba(9, 22, 42, 0.45));
      border: 1px solid rgba(148, 189, 255, 0.16);
      border-radius: 10px;
    }
    .mini {
      margin-top: 10px;
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 8px;
    }
    .mini .chip {
      background: rgba(9, 26, 49, 0.8);
      border: 1px solid rgba(148, 189, 255, 0.2);
      border-radius: 9px;
      padding: 9px;
      font-size: 12px;
    }
    .chip strong {
      display: block;
      margin-top: 4px;
      font-size: 16px;
      color: #e6f1ff;
    }

    .bar-list {
      display: grid;
      gap: 7px;
      margin-top: 8px;
    }
    .bar-row {
      display: grid;
      grid-template-columns: 84px 1fr 62px;
      gap: 8px;
      align-items: center;
      font-size: 12px;
    }
    .bar-track {
      height: 13px;
      background: var(--bar-bg);
      border-radius: 999px;
      overflow: hidden;
      border: 1px solid rgba(255, 255, 255, 0.09);
    }
    .bar-fill {
      width: 0;
      height: 100%;
      border-radius: 999px;
      background: linear-gradient(90deg, var(--accent), var(--accent2));
      transition: width 0.9s cubic-bezier(0.18, 0.8, 0.24, 1);
    }
    .bar-fill.alert {
      background: linear-gradient(90deg, #fb7185, #f97316);
    }
    .bar-fill.good {
      background: linear-gradient(90deg, #22c55e, #2dd4bf);
    }

    .heatmap-wrap {
      overflow: auto;
      border: 1px solid rgba(148, 189, 255, 0.16);
      border-radius: 10px;
      margin-top: 8px;
      max-height: 290px;
    }
    table {
      width: 100%;
      border-collapse: collapse;
      font-size: 12px;
      min-width: 540px;
    }
    th, td {
      padding: 8px 7px;
      border-bottom: 1px solid rgba(255, 255, 255, 0.08);
      text-align: center;
    }
    th:first-child, td:first-child {
      text-align: left;
      position: sticky;
      left: 0;
      background: #0d274a;
    }
    th {
      color: #b9d1ed;
      position: sticky;
      top: 0;
      background: #13345f;
      z-index: 2;
    }
    .heat-cell {
      border-radius: 7px;
      cursor: pointer;
      font-variant-numeric: tabular-nums;
      transition: transform 0.2s ease;
    }
    .heat-cell:hover {
      transform: translateY(-1px);
      outline: 1px solid rgba(255, 255, 255, 0.35);
    }

    .questions {
      margin: 0;
      padding-left: 18px;
      display: grid;
      gap: 8px;
    }
    .questions li {
      line-height: 1.45;
      color: #dce9fb;
      font-size: 13px;
    }
    .questions .priority {
      color: #ffe7af;
    }

    .issue-table-wrap {
      overflow: auto;
      border: 1px solid rgba(148, 189, 255, 0.16);
      border-radius: 10px;
      margin-top: 8px;
      max-height: 300px;
    }
    .issue-table td, .issue-table th {
      text-align: left;
      white-space: nowrap;
    }
    .issue-table td.msg {
      white-space: normal;
      min-width: 420px;
    }
    .badge {
      display: inline-block;
      border-radius: 999px;
      padding: 2px 7px;
      font-size: 11px;
      font-weight: 600;
      letter-spacing: 0.2px;
    }
    .badge.critical {
      color: #ffd0d7;
      background: rgba(251, 113, 133, 0.21);
    }
    .badge.warning {
      color: #ffe4a7;
      background: rgba(251, 191, 36, 0.23);
    }
    .badge.info {
      color: #bde7ff;
      background: rgba(96, 165, 250, 0.23);
    }

    .confidence {
      margin-top: 8px;
      border: 1px solid rgba(148, 189, 255, 0.2);
      border-radius: 11px;
      background: rgba(8, 23, 44, 0.65);
      padding: 10px;
    }
    .meter {
      margin-top: 8px;
      height: 14px;
      background: rgba(255, 255, 255, 0.12);
      border-radius: 999px;
      overflow: hidden;
    }
    .meter > div {
      height: 100%;
      width: 0;
      border-radius: 999px;
      background: linear-gradient(90deg, #fb7185, #fbbf24, #4ade80);
      transition: width 1s ease;
    }
    .confidence .meta {
      margin-top: 8px;
      color: #cde0f8;
      font-size: 12px;
      line-height: 1.45;
    }

    .future-note {
      margin-top: 10px;
      border: 1px dashed rgba(148, 189, 255, 0.34);
      border-radius: 10px;
      padding: 10px;
      color: #d7e7fb;
      font-size: 12px;
      line-height: 1.45;
      background: rgba(13, 31, 57, 0.5);
    }
    .footer {
      margin-top: 10px;
      color: var(--muted);
      font-size: 12px;
    }
    @media (max-width: 1240px) {
      .layout {
        grid-template-columns: 1fr;
      }
    }
    @media (max-width: 980px) {
      .kpis {
        grid-template-columns: 1fr 1fr;
      }
      .mini {
        grid-template-columns: 1fr;
      }
      .bar-row {
        grid-template-columns: 70px 1fr 58px;
      }
    }
  </style>
</head>
<body>
  <div class="wrap">
    <section class="hero">
      <h1>Trust Mid-Year Intelligence Board (Superpowered View)</h1>
      <p class="subtitle">This version is designed for trust/school leaders: animated trends, school-to-school comparisons, year-group diagnostics, and automatically generated challenge questions based on dips, jumps, and data quality risk.</p>
      <div class="kpis" id="kpiStrip"></div>
    </section>

    <div class="layout">
      <div class="stack">
        <section class="panel">
          <div class="panel-top">
            <div>
              <h2>Trust Trend Over Time</h2>
              <p>EYFS uses GLD. Years 1-6 use Combined ARE. Use this to spot trust-wide inflection points quickly.</p>
            </div>
          </div>
          <canvas id="trustTrendChart" class="chart"></canvas>
          <div class="mini" id="trustTrendInsights"></div>
        </section>

        <section class="panel">
          <div class="panel-top">
            <div>
              <h2>Year Group Comparison Across Schools</h2>
              <p>Choose a year to compare schools side-by-side and identify outliers for immediate support/challenge.</p>
            </div>
            <select id="yearPicker" class="control"></select>
          </div>
          <div id="yearBars" class="bar-list"></div>
          <div class="footer" id="yearSummary"></div>
        </section>

        <section class="panel">
          <div class="panel-top">
            <div>
              <h2>School Journey Explorer</h2>
              <p>Track each school's profile across EYFS to Year 6 and surface questions linked to dips and accelerations.</p>
            </div>
            <select id="schoolPicker" class="control"></select>
          </div>
          <canvas id="schoolTrendChart" class="chart"></canvas>
          <ul id="schoolQuestions" class="questions"></ul>
        </section>
      </div>

      <div class="stack">
        <section class="panel">
          <div class="panel-top">
            <div>
              <h2>Core Outcome Heatmap (School x Year)</h2>
              <p>Click any cell to pivot school/year context in the explorer.</p>
            </div>
          </div>
          <div class="heatmap-wrap">
            <table id="heatmapTable"></table>
          </div>
        </section>

        <section class="panel">
          <h2>What To Ask Schools Next</h2>
          <p>Auto-generated from trend breaks + integrity checks, prioritised for leadership conversations now.</p>
          <ol id="globalQuestions" class="questions"></ol>
          <div class="future-note">
            <strong>Cohort Tracking Extension:</strong> this file is a snapshot by year group, not pupil-level longitudinal records. If connected to historic DfE/MIS pupil-level data, this panel can switch to true cohort progression, moderation confidence checks, and intervention attribution by cohort.
          </div>
        </section>

        <section class="panel">
          <h2>Data Confidence & Risk Signals</h2>
          <p>Performance insight confidence should be interpreted alongside data reliability.</p>
          <div class="confidence">
            <div id="confidenceHeadline"></div>
            <div class="meter"><div id="confidenceBar"></div></div>
            <div class="meta" id="confidenceMeta"></div>
          </div>
          <div id="riskBreakdown" class="bar-list"></div>
        </section>

        <section class="panel">
          <h2>Priority Data Flags</h2>
          <p>These are likely to distort interpretation if unresolved.</p>
          <div class="issue-table-wrap">
            <table class="issue-table">
              <thead>
                <tr><th>Severity</th><th>Year</th><th>School</th><th>Issue</th></tr>
              </thead>
              <tbody id="issueRows"></tbody>
            </table>
          </div>
        </section>
      </div>
    </div>
  </div>

<script>
  const data = __PAYLOAD__;

  const years = data.trustOutcome.map(item => item.year);
  const yearLabels = years.map(year => year.replace("Year ", "Y"));
  const schools = data.heatmap.map(item => item.school);

  function pct(value) {
    if (value === null || value === undefined || Number.isNaN(value)) return "-";
    return `${(value * 100).toFixed(1)}%`;
  }
  function pp(value) {
    if (value === null || value === undefined || Number.isNaN(value)) return "-";
    return `${(value * 100).toFixed(1)}pp`;
  }
  function bySchool(school) {
    return data.heatmap.find(item => item.school === school) || null;
  }
  function coreLabel(year) {
    return year === "EYFS" ? "GLD" : "Combined ARE";
  }
  function clamp(value, min, max) {
    return Math.max(min, Math.min(max, value));
  }
  function makeCanvasHiDpi(canvas, widthPx, heightPx) {
    const ratio = window.devicePixelRatio || 1;
    canvas.width = Math.floor(widthPx * ratio);
    canvas.height = Math.floor(heightPx * ratio);
    canvas.style.width = `${widthPx}px`;
    canvas.style.height = `${heightPx}px`;
    const context = canvas.getContext("2d");
    context.setTransform(ratio, 0, 0, ratio, 0, 0);
    return context;
  }
  function drawAnimatedLineChart(canvasId, labels, values, options = {}) {
    const canvas = document.getElementById(canvasId);
    const width = canvas.clientWidth || 800;
    const height = canvas.clientHeight || 250;
    const context = makeCanvasHiDpi(canvas, width, height);

    const points = values
      .map((value, index) => ({ index, value, label: labels[index] }))
      .filter(point => point.value !== null && point.value !== undefined);
    if (!points.length) return;

    const minY = options.minY ?? Math.max(0, Math.min(...points.map(p => p.value)) - 0.05);
    const maxY = options.maxY ?? Math.min(1, Math.max(...points.map(p => p.value)) + 0.05);
    const left = 54;
    const right = 18;
    const top = 18;
    const bottom = 36;

    const mapX = idx => {
      if (labels.length <= 1) return left;
      return left + (idx * ((width - left - right) / (labels.length - 1)));
    };
    const mapY = value => top + (1 - ((value - minY) / (maxY - minY || 1))) * (height - top - bottom);

    const renderedPoints = points.map(point => ({ ...point, x: mapX(point.index), y: mapY(point.value) }));

    let startTime = null;
    const duration = 880;

    function frame(timestamp) {
      if (!startTime) startTime = timestamp;
      const t = clamp((timestamp - startTime) / duration, 0, 1);

      context.clearRect(0, 0, width, height);

      context.fillStyle = "rgba(10, 26, 49, 0.45)";
      context.fillRect(left, top, width - left - right, height - top - bottom);

      context.strokeStyle = "rgba(161, 193, 237, 0.2)";
      context.lineWidth = 1;
      [0.4, 0.5, 0.6, 0.7].forEach(tick => {
        const y = mapY(tick);
        context.beginPath();
        context.moveTo(left, y);
        context.lineTo(width - right, y);
        context.stroke();
        context.fillStyle = "#9fc0e6";
        context.font = "11px Avenir Next, Segoe UI, sans-serif";
        context.textAlign = "right";
        context.fillText(`${Math.round(tick * 100)}%`, left - 8, y + 4);
      });

      context.strokeStyle = "rgba(195, 220, 255, 0.45)";
      context.lineWidth = 1.2;
      context.beginPath();
      context.moveTo(left, top);
      context.lineTo(left, height - bottom);
      context.lineTo(width - right, height - bottom);
      context.stroke();

      const visibleCount = Math.max(1, Math.floor(renderedPoints.length * t));
      const visiblePoints = renderedPoints.slice(0, visibleCount);
      if (visiblePoints.length > 1) {
        context.strokeStyle = options.lineColor || "#2dd4bf";
        context.lineWidth = 3;
        context.beginPath();
        context.moveTo(visiblePoints[0].x, visiblePoints[0].y);
        for (let i = 1; i < visiblePoints.length; i += 1) {
          context.lineTo(visiblePoints[i].x, visiblePoints[i].y);
        }
        context.stroke();
      }

      renderedPoints.forEach(point => {
        context.fillStyle = "#eaf5ff";
        context.font = "11px Avenir Next, Segoe UI, sans-serif";
        context.textAlign = "center";
        context.fillText(point.label, point.x, height - 12);
      });

      visiblePoints.forEach(point => {
        context.fillStyle = options.dotColor || "#60a5fa";
        context.beginPath();
        context.arc(point.x, point.y, 4.2, 0, Math.PI * 2);
        context.fill();
        context.strokeStyle = "#f6fbff";
        context.lineWidth = 1;
        context.stroke();

        context.fillStyle = "#ebf6ff";
        context.font = "600 11px Avenir Next, Segoe UI, sans-serif";
        context.fillText(`${(point.value * 100).toFixed(1)}%`, point.x, point.y - 11);
      });

      if (t < 1) requestAnimationFrame(frame);
    }

    requestAnimationFrame(frame);
  }

  function buildKpis() {
    const strongest = data.strongestYear;
    const weakest = data.weakestYear;
    const widestGap = data.widestGap;
    const scoreTop = data.schoolScores[0];
    const scoreBottom = data.schoolScores[data.schoolScores.length - 1];

    const strip = document.getElementById("kpiStrip");
    const items = [
      {
        label: "Strongest Trust Year",
        value: strongest ? `${strongest.year}` : "-",
        sub: strongest ? `${pct(strongest.value)} (${coreLabel(strongest.year)})` : "-",
      },
      {
        label: "Weakest Trust Year",
        value: weakest ? `${weakest.year}` : "-",
        sub: weakest ? `${pct(weakest.value)} (${coreLabel(weakest.year)})` : "-",
      },
      {
        label: "Widest Deprivation Gap",
        value: widestGap ? widestGap.year : "-",
        sub: widestGap ? `${pp(widestGap.value)} (Non-FSM minus FSM)` : "-",
      },
      {
        label: "School Spread",
        value: `${scoreTop.school} to ${scoreBottom.school}`,
        sub: `${pct(scoreTop.value)} vs ${pct(scoreBottom.value)}`,
      },
      {
        label: "Data Risk Status",
        value: `${data.issueCounts.critical} critical`,
        sub: `${data.issueCounts.warning} warning / ${data.issueCounts.info} info`,
      },
    ];

    strip.innerHTML = items.map(item => `
      <article class="kpi">
        <div class="label">${item.label}</div>
        <div class="value">${item.value}</div>
        <div class="sub">${item.sub}</div>
      </article>
    `).join("");
  }

  function buildTrustTrend() {
    const values = data.trustOutcome.map(item => item.value);
    drawAnimatedLineChart("trustTrendChart", yearLabels, values, {
      minY: 0.35,
      maxY: 0.70,
      lineColor: "#2dd4bf",
      dotColor: "#60a5fa",
    });

    const deltas = [];
    for (let i = 1; i < data.trustOutcome.length; i += 1) {
      const previous = data.trustOutcome[i - 1];
      const current = data.trustOutcome[i];
      if (previous.value === null || current.value === null) continue;
      deltas.push({
        from: previous.year,
        to: current.year,
        delta: current.value - previous.value,
      });
    }
    const best = deltas.length ? deltas.reduce((a, b) => (b.delta > a.delta ? b : a)) : null;
    const worst = deltas.length ? deltas.reduce((a, b) => (b.delta < a.delta ? b : a)) : null;
    const gapValid = data.trustGap.filter(item => item.value !== null);
    const meanGap = gapValid.length
      ? gapValid.reduce((sum, item) => sum + item.value, 0) / gapValid.length
      : null;

    const insight = document.getElementById("trustTrendInsights");
    insight.innerHTML = `
      <div class="chip">Biggest rise<strong>${best ? `${best.from} -> ${best.to}: ${pp(best.delta)}` : "-"}</strong></div>
      <div class="chip">Biggest dip<strong>${worst ? `${worst.from} -> ${worst.to}: ${pp(worst.delta)}` : "-"}</strong></div>
      <div class="chip">Average deprivation gap<strong>${meanGap === null ? "-" : pp(meanGap)}</strong></div>
    `;
  }

  function animateBars(node) {
    const fills = node.querySelectorAll(".bar-fill");
    requestAnimationFrame(() => {
      fills.forEach(fill => {
        fill.style.width = fill.dataset.width || "0%";
      });
    });
  }

  function buildYearPicker() {
    const picker = document.getElementById("yearPicker");
    picker.innerHTML = years.map(year => `<option value="${year}">${year}</option>`).join("");
    picker.value = "Year 4";
    picker.addEventListener("change", renderYearComparison);
  }

  function renderYearComparison() {
    const year = document.getElementById("yearPicker").value;
    const rows = data.heatmap
      .map(item => ({ school: item.school, value: item[year] }))
      .filter(item => item.value !== null)
      .sort((a, b) => b.value - a.value);

    const trustAverage = rows.length
      ? rows.reduce((sum, item) => sum + item.value, 0) / rows.length
      : null;
    const maxValue = Math.max(...rows.map(item => item.value), 0.0001);

    const bars = document.getElementById("yearBars");
    bars.innerHTML = rows.map(item => `
      <div class="bar-row">
        <div>${item.school}</div>
        <div class="bar-track">
          <div class="bar-fill good" data-width="${(item.value / maxValue) * 100}%"></div>
        </div>
        <div style="text-align:right; font-variant-numeric: tabular-nums;">${pct(item.value)}</div>
      </div>
    `).join("");
    animateBars(bars);

    const best = rows[0];
    const worst = rows[rows.length - 1];
    const spread = best && worst ? best.value - worst.value : null;

    document.getElementById("yearSummary").textContent = trustAverage === null
      ? "No valid values for this year."
      : `Trust average ${pct(trustAverage)} (${coreLabel(year)}). Highest ${best.school} ${pct(best.value)}; lowest ${worst.school} ${pct(worst.value)}; spread ${pp(spread)}.`;
  }

  function buildSchoolPicker() {
    const picker = document.getElementById("schoolPicker");
    picker.innerHTML = schools.map(school => `<option value="${school}">${school}</option>`).join("");
    picker.value = data.schoolScores[0].school;
    picker.addEventListener("change", renderSchoolJourney);
  }

  function renderSchoolJourney() {
    const school = document.getElementById("schoolPicker").value;
    const record = bySchool(school);
    if (!record) return;

    const values = years.map(year => record[year]);
    drawAnimatedLineChart("schoolTrendChart", yearLabels, values, {
      minY: 0.28,
      maxY: 0.75,
      lineColor: "#60a5fa",
      dotColor: "#2dd4bf",
    });

    const transitions = [];
    for (let i = 1; i < years.length; i += 1) {
      const before = values[i - 1];
      const after = values[i];
      if (before === null || after === null) continue;
      transitions.push({ from: years[i - 1], to: years[i], delta: after - before });
    }
    const rise = transitions.length ? transitions.reduce((a, b) => (b.delta > a.delta ? b : a)) : null;
    const dip = transitions.length ? transitions.reduce((a, b) => (b.delta < a.delta ? b : a)) : null;
    const validValues = values.filter(v => v !== null);
    const variation = validValues.length ? Math.max(...validValues) - Math.min(...validValues) : null;

    const severeForSchool = data.issues.filter(issue =>
      issue.school === school && (issue.severity === "critical" || issue.severity === "warning")
    );

    const prompts = [];
    if (dip && dip.delta < -0.05) {
      prompts.push(
        `<span class="priority">Priority:</span> ${school} drops ${pp(dip.delta)} from ${dip.from} to ${dip.to}. What changed in cohort profile, staffing, curriculum sequence, or moderation approach during this transition?`
      );
    }
    if (rise && rise.delta > 0.06) {
      prompts.push(
        `${school} improves ${pp(rise.delta)} from ${rise.from} to ${rise.to}. Which specific strategies drove this, and how can they be codified for transfer into weaker year groups?`
      );
    }
    if (variation !== null && variation > 0.16) {
      prompts.push(
        `${school} shows a high internal spread (${pp(variation)}). Is curriculum and assessment implementation consistent across year teams?`
      );
    }
    if (severeForSchool.length >= 3) {
      prompts.push(
        `${school} has ${severeForSchool.length} serious data flags. Before acting on attainment narrative, should this school's returns be verified with phase leaders and trust assessment leads?`
      );
    }
    if (!prompts.length) {
      prompts.push(`No sharp volatility detected for ${school}. Next question: can this stable profile be used as a moderation benchmark for peer schools?`);
    }

    document.getElementById("schoolQuestions").innerHTML = prompts.map(item => `<li>${item}</li>`).join("");
  }

  function heatColor(value) {
    if (value === null || value === undefined) return "rgba(255,255,255,0.08)";
    const normal = clamp((value - 0.25) / 0.55, 0, 1);
    const red = Math.round(200 - normal * 125);
    const green = Math.round(48 + normal * 165);
    const blue = Math.round(60 + normal * 40);
    return `rgba(${red}, ${green}, ${blue}, 0.78)`;
  }

  function buildHeatmap() {
    const table = document.getElementById("heatmapTable");
    const head = `<thead><tr><th>School</th>${years.map(year => `<th>${year.replace("Year ", "Y")}</th>`).join("")}</tr></thead>`;
    const body = `<tbody>${data.heatmap.map(row => {
      const cells = years.map(year => {
        const value = row[year];
        const label = value === null ? "-" : `${(value * 100).toFixed(1)}%`;
        return `<td><span class="heat-cell" data-school="${row.school}" data-year="${year}" style="display:inline-block; min-width:66px; background:${heatColor(value)}">${label}</span></td>`;
      }).join("");
      return `<tr><td>${row.school}</td>${cells}</tr>`;
    }).join("")}</tbody>`;

    table.innerHTML = head + body;
    table.querySelectorAll(".heat-cell").forEach(cell => {
      cell.addEventListener("click", () => {
        const school = cell.getAttribute("data-school");
        const year = cell.getAttribute("data-year");
        document.getElementById("schoolPicker").value = school;
        renderSchoolJourney();
        document.getElementById("yearPicker").value = year;
        renderYearComparison();
      });
    });
  }

  function buildGlobalQuestions() {
    const questions = [];

    const trustTransitions = [];
    for (let i = 1; i < data.trustOutcome.length; i += 1) {
      const prev = data.trustOutcome[i - 1];
      const next = data.trustOutcome[i];
      if (prev.value === null || next.value === null) continue;
      trustTransitions.push({ from: prev.year, to: next.year, delta: next.value - prev.value });
    }
    const sharpestTrustDip = trustTransitions.length
      ? trustTransitions.reduce((a, b) => (b.delta < a.delta ? b : a))
      : null;
    if (sharpestTrustDip && sharpestTrustDip.delta < -0.04) {
      questions.push(
        `<span class="priority">Priority:</span> Trust-level dip of ${pp(sharpestTrustDip.delta)} from ${sharpestTrustDip.from} to ${sharpestTrustDip.to}. Which schools and subjects account for most of that movement?`
      );
    }

    const widestGap = data.widestGap;
    if (widestGap && widestGap.value > 0.15) {
      questions.push(
        `<span class="priority">Priority:</span> Deprivation gap peaks in ${widestGap.year} at ${pp(widestGap.value)}. What is your immediate disadvantaged pupil recovery plan in that phase?`
      );
    }

    const highRiskSchools = schools
      .map(school => ({
        school,
        count: data.issues.filter(issue =>
          issue.school === school && (issue.severity === "critical" || issue.severity === "warning")
        ).length,
      }))
      .filter(item => item.count > 0)
      .sort((a, b) => b.count - a.count);
    if (highRiskSchools.length) {
      const top = highRiskSchools[0];
      questions.push(
        `${top.school} has the highest volume of serious data flags (${top.count}). Should this school's submission be revalidated before strategic decisions are made?`
      );
    }

    const weightedIssues = data.issues.filter(issue => issue.check === "weighted_mismatch").length;
    if (weightedIssues > 0) {
      questions.push(
        `${weightedIssues} subgroup-weighted mismatches detected. Are subgroup definitions and all-pupil totals being entered consistently across schools?`
      );
    }

    const missingCore = data.issues.filter(issue => issue.check === "missing_core_metric").length;
    if (missingCore > 0) {
      questions.push(
        `${missingCore} missing core metric returns found. Can missing values be resolved this week so trust comparisons are complete?`
      );
    }

    if (!questions.length) {
      questions.push("No major trust-level alarm detected from this snapshot. Next step: compare these patterns against prior captures to test stability.");
    }
    document.getElementById("globalQuestions").innerHTML = questions.slice(0, 8).map(item => `<li>${item}</li>`).join("");
  }

  function buildConfidencePanel() {
    const critical = data.issueCounts.critical || 0;
    const warning = data.issueCounts.warning || 0;
    const info = data.issueCounts.info || 0;
    const confidence = clamp(100 - (critical * 3.8) - (warning * 1.5) - (info * 0.45), 8, 100);

    const headline = confidence >= 75
      ? "High confidence with targeted caveats"
      : confidence >= 55
        ? "Medium confidence; resolve key anomalies"
        : "Low confidence; verify data before action";

    document.getElementById("confidenceHeadline").innerHTML = `<strong>${headline}</strong> (${confidence.toFixed(0)}/100)`;
    document.getElementById("confidenceBar").style.width = `${confidence}%`;
    document.getElementById("confidenceMeta").innerHTML = `
      Critical issues: <strong>${critical}</strong><br/>
      Warning issues: <strong>${warning}</strong><br/>
      Informational issues: <strong>${info}</strong><br/>
      Recommendation: use confidence level alongside attainment trends in leadership discussions.
    `;

    const checkCounts = {};
    data.issues.forEach(issue => {
      const key = issue.check || "other";
      checkCounts[key] = (checkCounts[key] || 0) + 1;
    });
    const topChecks = Object.entries(checkCounts)
      .sort((a, b) => b[1] - a[1])
      .slice(0, 6);

    const max = Math.max(...topChecks.map(item => item[1]), 1);
    const list = document.getElementById("riskBreakdown");
    list.innerHTML = topChecks.map(([check, count]) => `
      <div class="bar-row">
        <div>${check}</div>
        <div class="bar-track"><div class="bar-fill alert" data-width="${(count / max) * 100}%"></div></div>
        <div style="text-align:right">${count}</div>
      </div>
    `).join("");
    animateBars(list);
  }

  function buildIssuesTable() {
    const rows = data.topIssues.map(issue => `
      <tr>
        <td><span class="badge ${issue.severity}">${issue.severity}</span></td>
        <td>${issue.year}</td>
        <td>${issue.school}</td>
        <td class="msg">${issue.message}</td>
      </tr>
    `).join("");
    document.getElementById("issueRows").innerHTML = rows;
  }

  function init() {
    buildKpis();
    buildTrustTrend();
    buildYearPicker();
    buildSchoolPicker();
    buildHeatmap();
    renderYearComparison();
    renderSchoolJourney();
    buildGlobalQuestions();
    buildConfidencePanel();
    buildIssuesTable();
  }

  window.addEventListener("resize", () => {
    buildTrustTrend();
    renderSchoolJourney();
  });
  init();
</script>
</body>
</html>
"""

    return html_template.replace("__PAYLOAD__", json.dumps(payload, ensure_ascii=False))


def build_findings_markdown(summary: dict[str, Any], issues: list[dict[str, Any]]) -> str:
    strong = summary["strongest_year"]
    weak = summary["weakest_year"]
    gap = summary["widest_gap"]

    critical = [issue for issue in issues if issue["severity"] == "critical"]
    warnings = [issue for issue in issues if issue["severity"] == "warning"]

    top_school = summary["school_scores"][0]
    bottom_school = summary["school_scores"][-1]

    lines: list[str] = []
    lines.append("# Trust Mid-Year Data Capture 2025/26: Analysis Report")
    lines.append("")
    lines.append("## Trust-level insights")
    lines.append(
        f"- Strongest trust year-level core outcome: **{strong['year']} ({pct(strong['value'])})**."
        if strong
        else "- Strongest trust year-level core outcome: n/a."
    )
    lines.append(
        f"- Weakest trust year-level core outcome: **{weak['year']} ({pct(weak['value'])})**."
        if weak
        else "- Weakest trust year-level core outcome: n/a."
    )
    lines.append(
        f"- Widest trust deprivation gap (Non-FSM minus FSM): **{gap['year']} ({gap['value'] * 100:.1f}pp)**."
        if gap
        else "- Widest trust deprivation gap: n/a."
    )
    lines.append(
        f"- Highest weighted school composite: **{top_school['school']} ({pct(top_school['value'])})**; lowest: **{bottom_school['school']} ({pct(bottom_school['value'])})**."
    )
    lines.append("")

    lines.append("## High-priority inconsistencies")
    if not critical and not warnings:
        lines.append("- No critical/warning data consistency issues detected.")
    else:
        for issue in (critical + warnings)[:18]:
            lines.append(f"- **{issue['severity'].upper()}** {issue['year']} / {issue['school']}: {issue['message']}")
    lines.append("")

    lines.append("## Suggested school follow-up questions")
    lines.append("- Are all **FSM counts** integer pupil counts in every tab (for example Year 1 HPS appears decimal)?")
    lines.append("- Can schools confirm whether **missing core metrics** are true missing returns or intentional exclusions?")
    lines.append("- Where subgroup weighted checks fail by >10pp, should FSM/Non-FSM rates be corrected or is the subgroup definition not a strict partition?")
    lines.append("- For rows where **GD exceeds ARE** (for example Year 4 GHPS FSM subgroup), is this data entry error or a field shift during copy/paste?")
    lines.append("- Should national comparators be consistently included for each year and all core metrics, with one numeric format (decimal proportions)?")
    lines.append("- For text-style entries (e.g. 'Above 25 64%'), should raw score and percentage be split into separate fields?")
    lines.append("")

    lines.append("## Data-engineering recommendations for app productisation")
    lines.append("- Store **long-format metric records** (`year`, `school`, `group`, `metric`, `value`) to simplify dynamic chart generation.")
    lines.append("- Enforce a schema with typed fields: integer denominators, proportion metrics (0-1), and optional raw score fields.")
    lines.append("- Add real-time validation rules for: denominator logic, GD<=ARE, combined<=subject minima, and subgroup weighted consistency.")
    lines.append("- Add confidence labels based on subgroup size thresholds to prevent over-interpreting very small cohorts.")
    lines.append("- Separate data quality diagnostics from attainment visuals so trust leaders can distinguish performance from data reliability.")

    return "\n".join(lines) + "\n"


def main() -> None:
    parser = argparse.ArgumentParser(description="Build trust mid-year analysis outputs")
    parser.add_argument(
        "--workbook",
        default="/Users/jarvis/Desktop/Trust mid year Data Capture 2025_26 (2).xlsx",
        help="Path to source workbook",
    )
    parser.add_argument(
        "--outdir",
        default="/Users/jarvis/dev/Schoolgle_Improvement/analysis_outputs/trust_midyear_2025_26",
        help="Output directory",
    )
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    records, raw_text_cells, national_rows = extract_records(workbook_path)
    issues = run_checks(records, raw_text_cells, national_rows)
    summary = build_summary(records, issues)

    write_csv(records, outdir / "trust_midyear_school_year_metrics.csv")

    with (outdir / "trust_midyear_analysis.json").open("w", encoding="utf-8") as handle:
        json.dump(
            {
                "workbook": str(workbook_path),
                "records": records,
                "raw_text_cells": raw_text_cells,
                "national_rows": national_rows,
                "issues": issues,
                "summary": summary,
            },
            handle,
            ensure_ascii=False,
            indent=2,
        )

    dashboard_html = build_dashboard_html(summary, issues)
    (outdir / "trust_midyear_dashboard.html").write_text(dashboard_html, encoding="utf-8")

    report_md = build_findings_markdown(summary, issues)
    (outdir / "trust_midyear_findings.md").write_text(report_md, encoding="utf-8")

    print(f"Wrote outputs to: {outdir}")
    print(f"Records: {len(records)}")
    print(f"Issues: {len(issues)}")


if __name__ == "__main__":
    main()
